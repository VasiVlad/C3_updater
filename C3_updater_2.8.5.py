# C3_updater.py — Финальная версия
# ✅ rename_files, rename_updated_files, check_duplicates
# ✅ CSV лог с заголовками и filename
# ✅ Чекбокс 95% включён по умолчанию
# ✅ Очистка окна логов

import os
import re
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import pandas as pd
from functools import partial
# Стили
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
bold_font = Font(bold=True, color="FF0000")
button_refs = []

LOG_FILENAME = "excel_updater_log.csv"
def init_log_file():
    with open(LOG_FILENAME, "w", encoding='cp1252') as f:
        f.write("Error Type;Tag;Row;Ref Name;Old Value;New Value;Sheet;Filename;Work Step Description;Unique Step\n")


def load_workstep_table():
    path = os.path.join(os.getcwd(), "WorkStep_Data.csv")
    try:
        return pd.read_csv(path, sep=';', encoding='cp1252')
    except Exception:
        return None

def get_step_description(df, sub_wc_code, ref_name):
    if df is None:
        return ""
    try:
        row = df[
            (df["Sub Work Class Code"].astype(str).str.strip() == str(sub_wc_code).strip()) &
            (df["Step Reference"].astype(str).str.strip() == str(ref_name).strip())
        ]
        return row["Work Step Description"].iloc[0] if not row.empty else ""
    except:
        return ""
try:
    test_df = pd.read_csv("WorkStep_Data.csv", sep=';', encoding='cp1252')
    print("✅ CSV loaded successfully")
    print(test_df.head(3))
except Exception as e:
    print("❌ CSV load failed:", e)

df_steps = load_workstep_table()


def log_to_file(error_type, tag, row, ref_name='', old_val='', new_val='', sheet='', filename='', step_description=''):
    unique_step = f"{tag}_{sheet}_{ref_name}"
    try:
        with open(LOG_FILENAME, "a", encoding="utf-8") as f:
            f.write(f"{error_type};{tag};{row};{ref_name};{old_val};{new_val};{sheet};{filename};{step_description};{unique_step}\n")

    except Exception as e:
        log(f"❌ Cannot write to log file ({LOG_FILENAME}): {e}")
        log(f"❌ Cannot write to log file ({LOG_FILENAME}): {e}")

def is_valid_sheet(sheetname):
    return bool(re.match(r'^\d{5}', sheetname))

def find_header_row(ws):
    for i in range(1, ws.max_row + 1):
        if ws[f"A{i}"].value == "WC":
            return i
    return None

def get_ref_columns(ws, header_row):
    for col in range(12, ws.max_column + 1):
        if ws.cell(row=header_row, column=col).value in (None, ""):
            return list(range(12, col))
    return []

def get_capped_columns(ws, header_row):
    capped = set()
    for col in range(12, ws.max_column + 1):
        text = ws.cell(row=header_row - 5, column=col).value
        if isinstance(text, str) and re.match(r'^[A-Z],[A-Z],[A-Z]$', text.strip()):
            capped.add(col)
    return capped

def sanitize_filename(filename):
    return re.sub(r'[\\/*?"<>|]', "", filename)

def log(msg):
    log_box.insert(tk.END, msg + "\n")
    log_box.see(tk.END)

def choose_folder(entry):
    folder = filedialog.askdirectory()
    if folder:
        entry.delete(0, tk.END)
        entry.insert(0, folder)
        update_contracts()

def get_selected(listbox):
    return [listbox.get(i) for i in listbox.curselection()]

def update_contracts(*_):
    root_easy = easy_path.get()
    if not os.path.isdir(root_easy): return
    contracts = sorted([c for c in os.listdir(root_easy) if os.path.isdir(os.path.join(root_easy, c))])
    contract_listbox.delete(0, tk.END)
    contract_listbox.insert(tk.END, "All")
    for c in contracts:
        contract_listbox.insert(tk.END, c)
    update_wcs()

def update_wcs(*_):
    root_easy = easy_path.get()
    selected = get_selected(contract_listbox)
    contracts = selected if "All" not in selected else os.listdir(root_easy)
    wcs = set()
    for contract in contracts:
        path = os.path.join(root_easy, contract)
        if os.path.isdir(path):
            for f in os.listdir(path):
                if f.endswith(".xlsx") and f[:5].isdigit():
                    wcs.add(f[:5])
    wc_listbox.delete(0, tk.END)
    wc_listbox.insert(tk.END, "All")
    for w in sorted(wcs):
        wc_listbox.insert(tk.END, w)

def rename_files(folder_path, suffix):
    renamed = 0
    for root, _, files in os.walk(folder_path):
        contract = os.path.basename(root)
        for filename in files:
            if filename.endswith(".xlsx"):
                match = re.search(r"\d{5}", filename)
                if match:
                    swc = match.group(0)
                    new_name = f"{swc}_{contract}_{suffix}.xlsx"
                    old_path = os.path.join(root, filename)
                    new_path = os.path.join(root, new_name)
                    if old_path != new_path:
                        os.rename(old_path, new_path)
                        log(f"Renamed: {filename} → {new_name}")
                        renamed += 1
    log(f"Total renamed in {suffix}: {renamed}")
    update_contracts()

def rename_updated_files():
    updated_root = updated_path.get()
    csv_path = os.path.join(os.getcwd(), "WorkClass_Data.csv")
    try:
        df = pd.read_csv(csv_path, sep=';')
        rename_dict = {
            str(row['Work Class Code']): sanitize_filename(' '.join(re.findall(r'\w+', row['Work Class Description'])[:5]))
            for _, row in df.iterrows()
        }
    except Exception as e:
        log(f"Error loading WorkClass_Data.csv: {e}")
        return

    updated_contracts = [f for f in os.listdir(updated_root) if os.path.isdir(os.path.join(updated_root, f))]
    for contract in updated_contracts:
        folder = os.path.join(updated_root, contract)
        for filename in os.listdir(folder):
            if re.match(r'^\d{5}_', filename):
                wc = filename.split('_')[0]
                if wc in rename_dict:
                    new_filename = f"{wc}_{rename_dict[wc]}_{contract}.xlsx"
                    old_path = os.path.join(folder, filename)
                    new_path = os.path.join(folder, new_filename)
                    os.rename(old_path, new_path)
                    log(f"Renamed: {filename} → {new_filename}")
                else:
                    log(f"⚠️ WC {wc} not found in dictionary")

def check_duplicates(ws, sheetname, filepath, files_with_duplicates):
    seen = {}
    duplicates = []
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        tag = row[4].value
        if tag is None:
            continue
        tag_str = str(tag).strip()
        if tag_str == "" or tag_str.lower() == "tag":
            continue
        if tag_str in seen:
            duplicates.append((tag_str, seen[tag_str], idx))
        else:
            seen[tag_str] = idx
    if duplicates:
        log(f"❗ Duplicates in {sheetname}: {len(duplicates)}")
        for tag, first, second in duplicates:
            log(f"    Tag '{tag}' in rows {first} and {second}")
            log_to_file("Duplicate", tag, f"{first},{second}", '', '', os.path.basename(filepath))
        files_with_duplicates.add(filepath)

def process_and_update():
    global button_refs
    for btn in button_refs:
        btn.destroy()
    button_refs.clear()
    log_box.delete('1.0', tk.END)
    init_log_file()

    selected_contracts = get_selected(contract_listbox)
    selected_wcs = get_selected(wc_listbox)

    if not selected_contracts or not selected_wcs:
        messagebox.showwarning("Missing selection", "Please select at least one contract and one WC.")
        return

    root_easy = easy_path.get()
    root_sub = sub_path.get()
    root_updated = updated_path.get()

    contracts = os.listdir(root_easy) if "All" in selected_contracts else selected_contracts
    files_with_duplicates = set()
    files_with_missing = set()

    for contract in contracts:
        easy_dir = os.path.join(root_easy, contract)
        sub_dir = os.path.join(root_sub, contract)
        updated_dir = os.path.join(root_updated, contract)
        os.makedirs(updated_dir, exist_ok=True)

        available_wcs = {f[:5] for f in os.listdir(easy_dir) if f.endswith(".xlsx") and f[:5].isdigit()}
        wcs = available_wcs if "All" in selected_wcs else selected_wcs

        for wc in wcs:
            easy_fp = os.path.join(easy_dir, f"{wc}_{contract}_easy.xlsx")
            sub_fp = os.path.join(sub_dir, f"{wc}_{contract}_sub.xlsx")
            updated_fp = os.path.join(updated_dir, f"{wc}_{contract}_updated.xlsx")

            if not os.path.exists(easy_fp) or not os.path.exists(sub_fp):
                log(f"❌ Missing files for {contract}/{wc}")
                continue

            try:
                wb_easy = load_workbook(easy_fp)
                wb_sub = load_workbook(sub_fp)

                for sheet in wb_easy.sheetnames:
                    if not is_valid_sheet(sheet) or sheet not in wb_sub.sheetnames:
                        continue

                    ws_easy = wb_easy[sheet]
                    ws_sub = wb_sub[sheet]

                    check_duplicates(ws_sub, sheet, sub_fp, files_with_duplicates)

                    header_row = find_header_row(ws_easy)
                    if not header_row:
                        continue

                    ref_cols = get_ref_columns(ws_easy, header_row)
                    capped_cols = get_capped_columns(ws_easy, header_row) if cap_ref_var.get() else set()

                    sub_data = {}
                    for row in ws_sub.iter_rows(min_row=header_row + 1):
                        tag = row[4].value
                        if tag:
                            sub_data[tag] = {col: row[col - 1].value for col in ref_cols}

                    for i, row in enumerate(ws_easy.iter_rows(min_row=header_row + 1), start=header_row + 1):
                        tag = row[4].value
                        if tag in sub_data:
                            for col in ref_cols:
                                cell = ws_easy.cell(row=i, column=col)
                                sub_val = sub_data[tag][col]
                                try:
                                    old = float(cell.value)
                                    new = float(sub_val)
                                    if col in capped_cols and new > 95:
                                        new = 95.0 if old < 95 else old
                                    if new > old:
                                        cell.value = new
                                        cell.font = bold_font
                                    elif new < old:
                                        cell.value = new
                                        cell.fill = yellow_fill
                                        msg = f"[↓] {wc}_{contract} | Row {i} | Tag: {tag} {old} → {new}"
                                        log(msg)
                                        ref_name = ws_easy.cell(row=header_row, column=col).value
                                        step_desc = get_step_description(df_steps, sheet, ref_name)
                                        log_to_file("Decreased", tag, i, ref_name, old, new, sheet, os.path.basename(updated_fp), step_desc)

                                except:
                                    cell.value = sub_val
                                    cell.fill = orange_fill
                                    if str(sub_val).strip().upper() not in ("X", "NA"):
                                        msg = f"⚠ Invalid value for Tag {tag} at row {i}, col {col}: '{sub_val}'"
                                        log(msg)
                                        ref_name = ws_easy.cell(row=header_row, column=col).value
                                        step_desc = get_step_description(df_steps, sheet, ref_name)
                                        log_to_file("Invalid", tag, i, ref_name, '', sub_val, sheet, os.path.basename(updated_fp), step_desc)

                        else:
                            for col in ref_cols:
                                cell = ws_easy.cell(row=i, column=col)
                                cell.value = 0
                                cell.fill = yellow_fill
                            msg = f"⚠ Missing tag {tag} in row {i} of {sheet}"
                            log(msg)
                            ref_name = ws_easy.cell(row=header_row, column=col).value
                            step_desc = get_step_description(df_steps, sheet, ref_name)
                            log_to_file("Missing", tag, i, ref_name, '', '', sheet, os.path.basename(updated_fp), step_desc)

                            files_with_missing.add(updated_fp)

                try:
                    wb_easy.save(updated_fp)
                    log(f"✔ Saved: {updated_fp}")
                except PermissionError:
                    log(f"❌ Cannot save {updated_fp} — file is open or write-protected")

            except Exception as e:
                log(f"❌ Error processing {wc}_{contract}: {e}")

    for f in sorted(files_with_duplicates):
        btn = tk.Button(root, text=f"Open DUP: {os.path.basename(f)}", command=partial(os.startfile, f))
        btn.grid()
        button_refs.append(btn)

    for f in sorted(files_with_missing):
        btn = tk.Button(root, text=f"Open MISSING: {os.path.basename(f)}", command=partial(os.startfile, f))
        btn.grid()
        button_refs.append(btn)
        
def open_selected_folders(base_path_entry):
    base_path = os.path.abspath(base_path_entry.get())
    selected = get_selected(contract_listbox)

    # Если ничего не выбрано или явно выбрано "All"
    if not selected or "All" in selected:
        if os.path.exists(base_path):
            os.startfile(base_path)
        else:
            log(f"❌ Path not found: {base_path}")
        return

    # Если выбраны конкретные контракты
    for contract in selected:
        contract_path = os.path.join(base_path, contract)
        if os.path.exists(contract_path):
            os.startfile(contract_path)
        else:
            log(f"❌ Contract folder not found: {contract_path}")




# GUI
root = tk.Tk()
root.title("Excel Updater")
root.geometry("1024x720")

tk.Label(root, text="Easy folder:").grid(row=0, column=0, sticky='e')
easy_path = tk.Entry(root, width=80)
easy_path.insert(0, "./Easy")
easy_path.grid(row=0, column=1)
tk.Button(root, text="Browse", command=lambda: choose_folder(easy_path)).grid(row=0, column=2)

tk.Label(root, text="Sub folder:").grid(row=1, column=0, sticky='e')
sub_path = tk.Entry(root, width=80)
sub_path.insert(0, "./Sub")
sub_path.grid(row=1, column=1)
tk.Button(root, text="Browse", command=lambda: choose_folder(sub_path)).grid(row=1, column=2)

tk.Label(root, text="Updated folder:").grid(row=2, column=0, sticky='e')
updated_path = tk.Entry(root, width=80)
updated_path.insert(0, "./Updated")
updated_path.grid(row=2, column=1)
tk.Button(root, text="Browse", command=lambda: choose_folder(updated_path)).grid(row=2, column=2)

tk.Button(root, text="Rename Easy", command=lambda: rename_files(easy_path.get(), "easy")).grid(row=3, column=0, pady=5)
tk.Button(root, text="Rename Sub", command=lambda: rename_files(sub_path.get(), "sub")).grid(row=3, column=1, pady=5)
tk.Button(root, text="Rename Updated", command=rename_updated_files).grid(row=3, column=2, pady=5)



# Лейбл
tk.Label(root, text="Select Contracts:").grid(row=4, column=0, sticky='ne')

# Общий фрейм для списков и кнопок
lists_frame = tk.Frame(root)
lists_frame.grid(row=4, column=1, columnspan=2, sticky='w', padx=10)

# Списки в колонке 0
contract_listbox = tk.Listbox(lists_frame, selectmode='multiple', exportselection=False, height=6)
contract_listbox.grid(row=0, column=0, sticky='w')

wc_listbox = tk.Listbox(lists_frame, selectmode='multiple', exportselection=False, height=6)
wc_listbox.grid(row=1, column=0, sticky='w')

contract_listbox.bind("<<ListboxSelect>>", update_wcs)

# Кнопки Open в колонке 1
open_frame = tk.Frame(lists_frame)
open_frame.grid(row=0, column=1, rowspan=2, padx=(10, 0), sticky='n')

tk.Button(open_frame, text="Open Easy", width=14, command=lambda: open_selected_folders(easy_path)).pack(pady=2)
tk.Button(open_frame, text="Open Sub", width=14, command=lambda: open_selected_folders(sub_path)).pack(pady=2)
tk.Button(open_frame, text="Open Updated", width=14, command=lambda: open_selected_folders(updated_path)).pack(pady=2)


# Чекбокс под списками
cap_ref_var = tk.BooleanVar(value=False)
tk.Checkbutton(root, text="Cap Ref values to 95%", variable=cap_ref_var).grid(row=6, column=1, sticky='w')


tk.Button(root, text="Update", command=process_and_update).grid(row=7, column=1, pady=10)

log_box = scrolledtext.ScrolledText(root, width=120, height=14)
log_box.grid(row=8, column=0, columnspan=3, padx=10, pady=10)

update_contracts()
root.mainloop()
